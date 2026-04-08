import sys
import glob 
import argparse
import pandas as pd
from pathlib import Path

# ── configurazione ──────────────────────────────────────────────────────────

ALLOWED_AREA_CODES = {
    "AT", "BA", "BE", "BG", "CH", "CZ", "DE",
    "DK", "EE", "ES", "FI", "FR", "GB",
    "GR", "HR", "HU", "IE", "IT", "LT", "LV", "MD",
    "ME", "NL", "NO", "PL", "PT", "RO", "RS",
    "SE", "SI", "SK"
}

# Colonne da leggere per ciascun file
USECOLS_GEN = [
    "DateTime(UTC)",
    "AreaMapCode",
    "GenerationForecast[MW]",
]

USECOLS_LOAD = [
    "DateTime(UTC)",
    "AreaMapCode",
    "TotalLoad[MW]",
]

DTYPE_GEN = {
    "AreaMapCode": "category",
    "GenerationForecast[MW]": "float32",
}

DTYPE_LOAD = {
    "AreaMapCode": "category",
    "TotalLoad[MW]": "float32",
}

# ────────────────────────────────────────────────────────────────────────────


def process_gen_file(path: str, chunksize: int) -> pd.DataFrame:
    """
    Legge un CSV di Day-Ahead Aggregated Generation in chunk.
    Restituisce DataFrame con colonne: Date, AreaMapCode, GenerationForecast[MW]
    aggregato a livello giornaliero (somma delle ore).
    """
    chunks = []
    for chunk in pd.read_csv(
        path,
        sep="\t",
        usecols=USECOLS_GEN,
        dtype=DTYPE_GEN,
        parse_dates=["DateTime(UTC)"],
        chunksize=chunksize,
        low_memory=False,
    ):
        # Filtra solo i paesi di interesse
        chunk = chunk[chunk["AreaMapCode"].isin(ALLOWED_AREA_CODES)].copy()
        if chunk.empty:
            continue

        # Estrai la data (normalizza a mezzanotte, scarta l'ora)
        chunk["Date"] = chunk["DateTime(UTC)"].dt.normalize()

        # Aggrega per giorno e paese: somma di tutte le ore del giorno
        agg = (
            chunk.groupby(["Date", "AreaMapCode"], observed=True)["GenerationForecast[MW]"]
            .sum()
            .reset_index()
        )
        chunks.append(agg)

    if not chunks:
        print(f"  [!] Nessun dato utile in {path}")
        return pd.DataFrame()

    return pd.concat(chunks, ignore_index=True)


def process_load_file(path: str, chunksize: int) -> pd.DataFrame:
    """
    Legge un CSV di Day-Ahead Total Load Forecast in chunk.
    Restituisce DataFrame con colonne: Date, AreaMapCode, TotalLoad[MW]
    aggregato a livello giornaliero (somma delle ore).
    """
    chunks = []
    for chunk in pd.read_csv(
        path,
        sep="\t",
        usecols=USECOLS_LOAD,
        dtype=DTYPE_LOAD,
        parse_dates=["DateTime(UTC)"],
        chunksize=chunksize,
        low_memory=False,
    ):
        # Filtra solo i paesi di interesse
        chunk = chunk[chunk["AreaMapCode"].isin(ALLOWED_AREA_CODES)].copy()
        if chunk.empty:
            continue

        # Estrai la data (normalizza a mezzanotte, scarta l'ora)
        chunk["Date"] = chunk["DateTime(UTC)"].dt.normalize()

        # Aggrega per giorno e paese
        agg = (
            chunk.groupby(["Date", "AreaMapCode"], observed=True)["TotalLoad[MW]"]
            .sum()
            .reset_index()
        )
        chunks.append(agg)

    if not chunks:
        print(f"  [!] Nessun dato utile in {path}")
        return pd.DataFrame()

    return pd.concat(chunks, ignore_index=True)


def build_pivot(df_gen: pd.DataFrame, df_load: pd.DataFrame) -> pd.DataFrame:
    """
    Riceve i DataFrame aggregati di generation e load,
    calcola il Reserve Margin per ogni (Date, Paese) e
    costruisce il pivot finale con una colonna per paese.

    Formula:
        RM = (GenerationForecast - LoadForecast) / LoadForecast

    Il risultato è adimensionale (es. 0.20 = 20% di margine).
    """

    # ── 1. Ri-aggrega (gestisce sovrapposizioni tra file diversi dello stesso mese)
    df_gen = (
        df_gen.groupby(["Date", "AreaMapCode"], observed=True)["GenerationForecast[MW]"]
        .sum()
        .reset_index()
    )
    df_load = (
        df_load.groupby(["Date", "AreaMapCode"], observed=True)["TotalLoad[MW]"]
        .sum()
        .reset_index()
    )

    # ── 2. Join su (Date, AreaMapCode)
    df = pd.merge(df_gen, df_load, on=["Date", "AreaMapCode"], how="inner")

    if df.empty:
        raise ValueError(
            "Il merge tra generation e load è vuoto: "
            "verifica che i file coprono lo stesso intervallo temporale e gli stessi paesi."
        )

    # ── 3. Calcolo Reserve Margin
    # Evitiamo divisione per zero: se LoadForecast == 0 il risultato è NaN
    df["ReserveMargin"] = (
        (df["GenerationForecast[MW]"] - df["TotalLoad[MW]"])
        / df["TotalLoad[MW]"].replace(0, float("nan"))
    )

    # ── 4. Costruisci il pivot: righe = Date, colonne = "AT_RM", "BE_RM", ...
    df["Column"] = df["AreaMapCode"].astype(str) + "_RM"

    pivot = df.pivot_table(
        index="Date",
        columns="Column",
        values="ReserveMargin",
        aggfunc="mean",   # media nel caso (raro) di duplicati residui
    )
    pivot.index.name = "Date"
    pivot.columns.name = None

    # Ordina colonne alfabeticamente (coerente con gen_data)
    pivot = pivot.sort_index(axis=1)
    return pivot


def save_excel(pivot: pd.DataFrame, output_path: str) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Reserve Margin")
        ws = writer.sheets["Reserve Margin"]

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", start_color="2D6A9F", end_color="2D6A9F")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell_font   = Font(name="Arial", size=10)

        # Formatta intestazione
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Larghezza colonne
        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        # Formato percentuale per i valori di Reserve Margin
        for row in ws.iter_rows(min_row=2):
            row[0].font = cell_font
            for cell in row[1:]:
                cell.font = cell_font
                cell.number_format = "0.00%"   # es. 0.20 → "20.00%"

        ws.freeze_panes = "B2"

    print(f"\n✅  Salvato: {output_path}")
    
def expand_paths(patterns: list[str]) -> list[str]:
    expanded = []
    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            expanded.extend(sorted(matches))
        else:
            expanded.append(pattern)
    return expanded


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description=(
            "Calcola il Reserve Margin giornaliero per paese da file ENTSO-E.\n"
            "Formula: RM = (GenerationForecast - LoadForecast) / LoadForecast\n\n"
            "Uso:\n"
            "  python reserve_load.py \\\n"
            "    --gen  2024_01_DayAheadAggregatedGeneration_14.1.C_r3.csv \\\n"
            "           2024_02_DayAheadAggregatedGeneration_14.1.C_r3.csv \\\n"
            "    --load 2024_01_DayAheadTotalLoadForecast_6.1.B_r3.csv \\\n"
            "           2024_02_DayAheadTotalLoadForecast_6.1.B_r3.csv"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--gen", nargs="+", required=True,
        help="Uno o più file CSV DayAheadAggregatedGeneration"
    )
    parser.add_argument(
        "--load", nargs="+", required=True,
        help="Uno o più file CSV DayAheadTotalLoadForecast"
    )
    parser.add_argument("-o", "--output", default="reserve_margin_data.xlsx")
    parser.add_argument("--no-excel", action="store_true")
    parser.add_argument("--chunksize", type=int, default=200_000)
    args = parser.parse_args()

    # ── Leggi tutti i file generation
    gen_files  = expand_paths(args.gen)
    load_files = expand_paths(args.load)
    all_gen = []
    for path in gen_files:
        print(f"→ Gen read:  {path}")
        df = process_gen_file(path, args.chunksize)
        if not df.empty:
            all_gen.append(df)

    # ── Leggi tutti i file load
    all_load = []
    for path in load_files:
        print(f"→ Load read: {path}")
        df = process_load_file(path, args.chunksize)
        if not df.empty:
            all_load.append(df)

    if not all_gen or not all_load:
        print("Dati insufficienti: verifica i file di input.")
        sys.exit(1)

    df_gen  = pd.concat(all_gen,  ignore_index=True)
    df_load = pd.concat(all_load, ignore_index=True)

    pivot = build_pivot(df_gen, df_load)

    print(f"\nPivot shape: {pivot.shape}  ({pivot.index.min().date()} → {pivot.index.max().date()})")
    print(pivot.head(3).to_string())

    if not args.no_excel:
        save_excel(pivot, args.output)


if __name__ == "__main__":
    main()