import sys
import argparse
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ALLOWED_AREA_CODES = {
    "AT", "BA", "BE", "BG", "CH", "CY", "CZ", "DE",
    "DK", "EE", "ES", "FI", "FR", "GB", "GE",
    "GR", "HR", "HU", "IE", "IT", "LT", "LU", "LV", "MD",
    "ME", "NL", "NO", "PL", "PT", "RO", "RS",
    "SE", "SI", "SK"
}

DTYPE = {
    "AreaMapCode": "category",
}


def process_file(path: str, chunksize: int) -> pd.DataFrame:
    # rileva separatore e nomi colonne reali
    sample = pd.read_csv(path, nrows=1, sep=None, engine="python")
    sep = "\t" if "DateTime(UTC)" in sample.columns else ","

    usecols = ["DateTime(UTC)", "AreaMapCode", "TotalLoad[MW]"]

    chunks = []
    for chunk in pd.read_csv(
        path,
        sep=sep,
        usecols=usecols,
        dtype=DTYPE,
        parse_dates=["DateTime(UTC)"],
        chunksize=chunksize,
        low_memory=False,
    ):
        mask = chunk["AreaMapCode"].isin(ALLOWED_AREA_CODES)
        chunk = chunk[mask].copy()
        if chunk.empty:
            continue

        chunk["Date"] = chunk["DateTime(UTC)"].dt.normalize()

        agg = (
            chunk.groupby(["Date", "AreaMapCode"], observed=True)["TotalLoad[MW]"]
            .mean()
            .reset_index()
        )
        chunks.append(agg)

    if not chunks:
        print(f"  [!] Nessun dato utile in {path}")
        return pd.DataFrame()

    return pd.concat(chunks, ignore_index=True)


def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    df = (
        df.groupby(["Date", "AreaMapCode"], observed=True)["TotalLoad[MW]"]
        .mean()
        .reset_index()
    )
    df["Column"] = df["AreaMapCode"].astype(str) + "_Load"

    pivot = df.pivot_table(
        index="Date",
        columns="Column",
        values="TotalLoad[MW]",
        aggfunc="mean",
        fill_value=0,
    )
    pivot.index.name = "Date"
    pivot.columns.name = None
    return pivot.sort_index(axis=1)


def save_excel(pivot: pd.DataFrame, output_path: str) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Total Load")
        ws = writer.sheets["Total Load"]

        header_fill = PatternFill("solid", start_color="2D6A9F", end_color="2D6A9F")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell_font = Font(name="Arial", size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        for row in ws.iter_rows(min_row=2):
            row[0].font = cell_font
            for cell in row[1:]:
                cell.font = cell_font
                cell.number_format = "#,##0.0"

        ws.freeze_panes = "B2"

    print(f"\n✅  Salvato: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Pivota CSV TotalLoad ENTSO-E")
    parser.add_argument("files", nargs="+", help="Uno o più file CSV")
    parser.add_argument("-o", "--output", default="total_load_pivot.xlsx")
    parser.add_argument("--chunksize", type=int, default=200_000)
    args = parser.parse_args()

    all_agg = []
    for path in args.files:
        print(f"→ Leggo: {path}")
        agg = process_file(path, args.chunksize)
        if not agg.empty:
            all_agg.append(agg)

    if not all_agg:
        print("Nessun dato trovato.")
        sys.exit(1)

    combined = pd.concat(all_agg, ignore_index=True)
    pivot = build_pivot(combined)

    print(f"\nPivot shape: {pivot.shape}  ({pivot.index.min().date()} → {pivot.index.max().date()})")
    print(pivot.head(3).to_string())

    save_excel(pivot, args.output)


if __name__ == "__main__":
    main()