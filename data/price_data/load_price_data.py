import sys
import argparse
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------
# ENTSO-E bidding zone -> ISO-2 country mapping
# Countries with a single nationwide zone use their 2-letter code directly.
# Multi-zone countries are aggregated (simple mean across zones per day).
# ---------------------------------------------------------------
ZONE_TO_COUNTRY = {
    # Single-zone countries (identity mapping)
    "AT": "AT", "BE": "BE", "BG": "BG", "CH": "CH", "CZ": "CZ",
    "EE": "EE", "ES": "ES", "FI": "FI", "FR": "FR", "GB": "GB",
    "GR": "GR", "HR": "HR", "HU": "HU", "LT": "LT", "LV": "LV",
    "MD": "MD", "ME": "ME", "NL": "NL", "PL": "PL", "PT": "PT",
    "RO": "RO", "RS": "RS", "SI": "SI", "SK": "SK",
    "BA": "BA",

    # Germany-Luxembourg joint zone 
    "DE_LU":    "DE",

    # Italy 
    
    "IT-Calabria": "IT", "IT-CNORTH": "IT", "IT-CSOUTH": "IT",
    "IT-NORTH": "IT", "IT-SACOAC": "IT", "IT-SACODC": "IT", 
    "IT-Sardinia": "IT", "IT-Sicily": "IT", "IT-SOUTH": "IT",

    # Norway 
    "NO1": "NO", "NO2": "NO", "NO3": "NO", "NO4": "NO", "NO5": "NO",
    "NO2NSL": "NO",

    # Sweden 
    "SE1": "SE", "SE2": "SE", "SE3": "SE", "SE4": "SE",

    # Denmark 
    "DK1": "DK", "DK2": "DK",

    # Ireland - Single Electricity Market (Ireland + Northern Ireland)
    "IE_SEM": "IE",
}

ALLOWED_MAP_CODES = set(ZONE_TO_COUNTRY.keys())

DTYPE = {
    "MapCode": "object",
    "Currency": "category",
}


def process_file(path: str, chunksize: int) -> pd.DataFrame:
    # Auto-detect separator
    sample = pd.read_csv(path, nrows=1, sep=None, engine="python")
    sep = "\t" if "DateTime(UTC)" in sample.columns else ","

    usecols = ["DateTime(UTC)", "MapCode", "Price[Currency/MWh]"]

    chunks = []
    for chunk in pd.read_csv(
        path,
        sep=sep,
        usecols=usecols,
        dtype=DTYPE,
        parse_dates=["DateTime(UTC)"],
        chunksize=chunksize,
        low_memory=False,
    ):
        
        # Keep only rows for zones we care about
        mask = chunk["MapCode"].isin(ALLOWED_MAP_CODES)
        chunk = chunk[mask].copy()
        if chunk.empty:
            continue

        # Daily normalisation
        chunk["Date"] = chunk["DateTime(UTC)"].dt.normalize()

        # Map each zone to its country BEFORE aggregation
        # .astype(str) is needed because MapCode is categorical
        # Modifica questa riga nel tuo script:
        chunk["Country"] = chunk["MapCode"].astype(str).str.strip().map(ZONE_TO_COUNTRY)

        # Step 1 - hourly to daily within each zone (mean)
        daily_zone = (
            chunk.groupby(["Date", "Country", "MapCode"], observed=True)[
                "Price[Currency/MWh]"
            ]
            .mean()
            .reset_index()
        )

        # Step 2 - across zones within each country (mean)
        # For single-zone countries this is a no-op.
        daily_country = (
            daily_zone.groupby(["Date", "Country"], observed=True)[
                "Price[Currency/MWh]"
            ]
            .mean()
            .reset_index()
        )
        chunks.append(daily_country)

    if not chunks:
        print(f"  [!] No useful data found in {path}")
        return pd.DataFrame()

    return pd.concat(chunks, ignore_index=True)


def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    # Final aggregation to merge duplicates coming from multiple input files
    df = (
        df.groupby(["Date", "Country"], observed=True)["Price[Currency/MWh]"]
        .mean()
        .reset_index()
    )
    df["Column"] = df["Country"].astype(str) + "_Price"

    # NOTE: fill_value removed - missing days stay as NaN for proper downstream imputation
    pivot = df.pivot_table(
        index="Date",
        columns="Column",
        values="Price[Currency/MWh]",
        aggfunc="mean",
    )
    pivot.index.name = "Date"
    pivot.columns.name = None
    return pivot.sort_index(axis=1)


def save_excel(pivot: pd.DataFrame, output_path: str) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Day-Ahead Prices")
        ws = writer.sheets["Day-Ahead Prices"]

        header_fill = PatternFill("solid", start_color="2D6A9F", end_color="2D6A9F")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell_font = Font(name="Arial", size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        for row in ws.iter_rows(min_row=2):
            row[0].font = cell_font
            for cell in row[1:]:
                cell.font = cell_font
                cell.number_format = "#,##0.00"

        ws.freeze_panes = "B2"

    print(f"\nSaved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Pivot CSV Day-Ahead Prices ENTSO-E (zone-aware)")
    parser.add_argument("files", nargs="+", help="One or more CSV files")
    parser.add_argument("-o", "--output", default="prices_pivot.xlsx")
    parser.add_argument("--chunksize", type=int, default=200_000)
    args = parser.parse_args()

    all_agg = []
    for path in args.files:
        print(f"-> Reading: {path}")
        agg = process_file(path, args.chunksize)
        if not agg.empty:
            all_agg.append(agg)

    if not all_agg:
        print("No data found.")
        sys.exit(1)

    combined = pd.concat(all_agg, ignore_index=True)
    pivot = build_pivot(combined)

    print(f"\nPivot shape: {pivot.shape} ({pivot.index.min().date()} -> {pivot.index.max().date()})")
    print(f"Countries recovered: {sorted(c.replace('_Price','') for c in pivot.columns)}")
    print("\nFirst rows:")
    print(pivot.head(3).to_string())

    save_excel(pivot, args.output)


if __name__ == "__main__":
    main()