import sys
import argparse
import pandas as pd
from pathlib import Path

# ── conf ──────────────────────────────────────────────────────────
PRODUCTION_MAP = {
    "Solar": "Solar",
    "Wind Onshore": "Wind",
    "Wind Offshore": "Wind",
    "Hydro Run-of-river and poundage": "Hydro",
}


ALLOWED_AREA_CODES = {
    "AT", "BA", "BE", "BG", "CH", "CZ", "DE",
    "DK", "EE", "ES", "FI", "FR", "GB",
    "GR", "HR", "HU", "IE", "IT", "LT", "LV", "MD",
    "ME",  "NL", "NO", "PL", "PT", "RO", "RS",
    "SE", "SI", "SK"
}

USECOLS = [
    "DateTime(UTC)",
    "AreaMapCode",
    "ProductionType",
    "ActualGenerationOutput[MW]",
]

DTYPE = {
    "AreaMapCode": "category",
    "ProductionType": "category",
    "ActualGenerationOutput[MW]": "float32",
}
# ────────────────────────────────────────────────────────────────────────────


def process_file(path: str, chunksize: int) -> pd.DataFrame:
    chunks = []
    for chunk in pd.read_csv(
        path,
        sep="\t",          
        usecols=USECOLS,
        dtype=DTYPE,
        parse_dates=["DateTime(UTC)"],
        chunksize=chunksize,
        low_memory=False,
    ):
        # filter ProductionType of interest
        mask = chunk["ProductionType"].isin(PRODUCTION_MAP)
        mask = mask & chunk["AreaMapCode"].isin(ALLOWED_AREA_CODES)
        chunk = chunk[mask].copy()
        if chunk.empty:
            continue

        # map → aggregate type 
        chunk["EnergyType"] = chunk["ProductionType"].map(PRODUCTION_MAP)

        # exctract date
        chunk["Date"] = chunk["DateTime(UTC)"].dt.normalize()

        # agggregate per day / country / type
        agg = (
            chunk.groupby(["Date", "AreaMapCode", "EnergyType"], observed=True)[
                "ActualGenerationOutput[MW]"
            ]
            .sum()
            .reset_index()
        )
        chunks.append(agg)

    if not chunks:
        print(f"  [!] No data in {path}")
        return pd.DataFrame()

    return pd.concat(chunks, ignore_index=True)


def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    # sum 
    df = (
        df.groupby(["Date", "AreaMapCode", "EnergyType"], observed=True)[
            "ActualGenerationOutput[MW]"
        ]
        .sum()
        .reset_index()
    )

    # combined col creation
    df["Column"] = df["AreaMapCode"].astype(str) + "_" + df["EnergyType"]

    pivot = df.pivot_table(
        index="Date",
        columns="Column",
        values="ActualGenerationOutput[MW]",
        aggfunc="sum",
        fill_value=0,
    )
    pivot.index.name = "Date"
    pivot.columns.name = None

    # alphabetical sort
    pivot = pivot.sort_index(axis=1)
    return pivot


def save_excel(pivot: pd.DataFrame, output_path: str) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Energy Pivot")
        ws = writer.sheets["Energy Pivot"]

        # format date col
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", start_color="2D6A9F", end_color="2D6A9F")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell_font   = Font(name="Arial", size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # col width
        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # format number for dati MW
        for row in ws.iter_rows(min_row=2):
            row[0].font = cell_font  # data
            for cell in row[1:]:
                cell.font = cell_font
                cell.number_format = "#,##0.0"

        # freeze header
        ws.freeze_panes = "B2"

    print(f"\n✅  Saved: {output_path}")


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Pivot CSV ENTSO-E for Solar/Wind/Hydro")
    parser.add_argument("files", nargs="+", help="One or more file CSV")
    parser.add_argument("-o", "--output", default="gen_data.xlsx")
    parser.add_argument("--no-excel", action="store_true")
    parser.add_argument("--chunksize", type=int, default=200_000)
    args = parser.parse_args()

    all_agg = []
    for path in args.files:
        print(f"→ Read: {path}")
        agg = process_file(path, args.chunksize)
        if not agg.empty:
            all_agg.append(agg)

    if not all_agg:
        print("Nessun dato trovato. Controlla i file e il separatore (tab).")
        sys.exit(1)

    combined = pd.concat(all_agg, ignore_index=True)
    pivot = build_pivot(combined)

    print(f"\nPivot shape: {pivot.shape}  ({pivot.index.min().date()} → {pivot.index.max().date()})")
    print(pivot.head(3).to_string())

    if not args.no_excel:
        save_excel(pivot, args.output)


if __name__ == "__main__":
    main()