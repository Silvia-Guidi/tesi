
"""
excel_export.py
===============

Export the full network analysis results to a single multi-sheet Excel
file. Each sheet contains a different angle of the analysis, all in
tabular form, ready for thesis reporting and further inspection.

Sheets produced
---------------
  1. README                   Glossary, conventions, how to read.
  2. G0_summary               One row: graph-level metrics with 90% CI.
  3. G0_nodes                 28 rows: per-country G0 metrics with 90% CI.
  4. G0_edges                 ~78 rows: edge probability, ci_lower, n_eff per
                              admissible directed edge.
  5. Phi_nodes                28 rows wide: all node metrics for lag1, lag7, agg.
  6. Phi_graph_level          3 rows (lag1, lag7, agg): total weight, density,
                              gini, mean weight.
  7. Gamma_exposure           28 rows: own / cross / wind / solar exposure.
  8. Gamma_pervasiveness      53 rows: per-exogenous pervasiveness ranked.
  9. Gamma_country_net_nodes  28 rows: collapsed-network metrics for
                              wind / solar / combined.
 10. Gamma_country_net_graph  3 rows: graph-level metrics on the collapsed nets.
 11. Cross_comparison         28 rows: each country's rank in G0, Phi, Gamma.

Conventions
-----------
- Header row in bold + light fill, frozen top row, auto-width columns.
- Numbers formatted to 4 decimals (3 for shares/percentages already in %).
- Rows are pre-sorted in the most meaningful way per sheet (typically
  descending by the most informative column).
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from network_load        import ChainOutput
from edge_analysis         import EdgeAnalysisBundle
from network_metrics       import MetricsBundle as G0MetricsBundle
from Phi_analysis   import WeightedMetricsBundle
from Gamma_analysis import GammaAnalysisResults


# ============================================================
# STYLE CONSTANTS
# ============================================================

HEADER_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL     = PatternFill("solid", start_color="3D5A80")
HEADER_ALIGN    = Alignment(horizontal="center", vertical="center")
BODY_FONT       = Font(name="Arial", size=10)
TITLE_FONT      = Font(name="Arial", bold=True, size=12, color="3D5A80")
NOTE_FONT       = Font(name="Arial", italic=True, size=9, color="555555")

THIN_BORDER = Border(
    left  =Side(style="thin", color="BBBBBB"),
    right =Side(style="thin", color="BBBBBB"),
    top   =Side(style="thin", color="BBBBBB"),
    bottom=Side(style="thin", color="BBBBBB"),
)

NUM_FMT_4 = "0.0000"
NUM_FMT_3 = "0.000"
NUM_FMT_PCT = "0.0%"
NUM_FMT_INT = "0"


# ============================================================
# LOW-LEVEL HELPERS
# ============================================================

def _write_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Write a styled header row at the given 1-based row index."""
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = THIN_BORDER


def _write_row(ws: Worksheet,
               row: int,
               values: list,
               number_formats: Optional[list[str]] = None) -> None:
    """Write a data row, optionally applying per-column number formats."""
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font   = BODY_FONT
        cell.border = THIN_BORDER
        if number_formats is not None and c - 1 < len(number_formats):
            fmt = number_formats[c - 1]
            if fmt:
                cell.number_format = fmt


def _autosize_columns(ws: Worksheet, min_width: int = 10) -> None:
    """Approximate auto-width: use the longest string in each column."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _freeze_header(ws: Worksheet, row: int = 2) -> None:
    """Freeze panes so the header row stays visible while scrolling."""
    ws.freeze_panes = f"A{row}"


def _rank_descending(values: np.ndarray) -> np.ndarray:
    """
    Return the 1-based rank of each value, where the largest gets rank 1.
    Ties are broken by stable order (first encountered wins).
    """
    n = len(values)
    order = np.argsort(-values, kind="stable")    # descending
    ranks = np.empty(n, dtype=int)
    for r, idx in enumerate(order):
        ranks[idx] = r + 1
    return ranks


# ============================================================
# SHEET 1 — README
# ============================================================

def _build_readme(ws: Worksheet,
                  chain: ChainOutput,
                  alpha: float = 0.10) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    rows = [
        ("BGVAR Network Analysis — results workbook", None, "title"),
        ("", None, None),
        ("Chain configuration", None, "title"),
        ("Countries (ny)",     chain.ny,               "num"),
        ("Exogenous (nx)",     chain.nx,               "num"),
        ("Lags",               str(chain.selected_lags), "txt"),
        ("Posterior samples",  chain.n_keep,           "num"),
        ("Credible interval",  f"{int(100*(1-alpha))}% (one-sided where applicable)", "txt"),
        ("", None, None),
        ("Convention on graph orientation", None, "title"),
        ("G[i, j] = 1", "node j influences node i  (j is the parent, i is the response)", "txt"),
        ("in-degree of i",  "number of parents (sources) of i  =  sum_j G[i, j]", "txt"),
        ("out-degree of j", "number of children (targets) of j  =  sum_i G[i, j]", "txt"),
        ("", None, None),
        ("How to read each sheet", None, "title"),
        ("G0_summary",    "Graph-level metrics for G0 (contemporaneous DAG) with 90% CI.", "txt"),
        ("G0_nodes",      "Per-country metrics for G0 with 90% CI. Sorted by total_degree desc.", "txt"),
        ("G0_edges",      "Posterior edge probability for each admissible directed edge of G0.", "txt"),
        ("Phi_nodes",     "Weighted metrics on |Phi| posterior mean. Three blocks: lag1, lag7, agg.", "txt"),
        ("Phi_graph_level","Graph-level summaries for each Phi slice.", "txt"),
        ("Gamma_exposure","Per-country exogenous exposure decomposed in own/cross x wind/solar.", "txt"),
        ("Gamma_pervasiveness", "Per-exogenous pervasiveness (column out-strength in G_Gamma).", "txt"),
        ("Gamma_country_net_nodes", "Metrics on the 28x28 country-collapsed network (Gamma).", "txt"),
        ("Gamma_country_net_graph", "Graph-level metrics on the collapsed network.", "txt"),
        ("Cross_comparison", "Per-country rank in G0 / Phi / Gamma for direct comparison.", "txt"),
        ("", None, None),
        ("Important caveats", None, "title"),
        ("G0 is a DAG by construction", "Reciprocity = 0 and # SCC = ny are structural, not findings.", "txt"),
        ("Phi/Gamma weighted networks", "Built on POSTERIOR MEAN coefficients (ESS > 1000); the binary "
         "G_Phi/G_Gamma samples have ESS ~50 and were intentionally not used "
         "for centrality computations.", "txt"),
        ("Weights = |posterior_mean|", "Sign of Phi/Gamma indicates direction of effect (deficit vs excess); "
         "magnitude is what matters for influence.", "txt"),
    ]

    r = 1
    for label, value, kind in rows:
        if kind == "title":
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = TITLE_FONT
        elif kind == "num":
            ws.cell(row=r, column=1, value=label).font = BODY_FONT
            ws.cell(row=r, column=2, value=value).font = BODY_FONT
        elif kind == "txt":
            ws.cell(row=r, column=1, value=label).font = BODY_FONT
            cell_v = ws.cell(row=r, column=2, value=value)
            cell_v.font = BODY_FONT
            cell_v.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


# ============================================================
# SHEET 2 — G0 GRAPH-LEVEL SUMMARY
# ============================================================

def _build_G0_summary(ws: Worksheet, g0: G0MetricsBundle) -> None:
    headers = ["metric", "mean", "median", "ci_low", "ci_high", "note"]
    _write_header(ws, 1, headers)

    rows = [
        ("density",      g0.density,      "fraction of admissible edges active"),
        ("transitivity", g0.transitivity, "directed clustering coefficient"),
        ("reciprocity",  g0.reciprocity,  "0 by construction (DAG)"),
        ("n_strong_cc",  g0.n_strong_cc,  "ny by construction (DAG)"),
    ]
    fmts = [None, NUM_FMT_4, NUM_FMT_4, NUM_FMT_4, NUM_FMT_4, None]
    for r, (name, gm, note) in enumerate(rows, start=2):
        _write_row(ws, r, [name, gm.mean, gm.median, gm.ci_low, gm.ci_high, note], fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 3 — G0 NODE METRICS
# ============================================================

def _build_G0_nodes(ws: Worksheet, g0: G0MetricsBundle) -> None:
    headers = [
        "country",
        "in_mean",   "in_low",   "in_high",
        "out_mean",  "out_low",  "out_high",
        "total_mean","total_low","total_high",
        "eigen_mean","eigen_low","eigen_high",
        "between_mean","between_low","between_high",
    ]
    _write_header(ws, 1, headers)

    # Sort by total_degree mean, descending
    order = np.argsort(-g0.total_degree.mean, kind="stable")

    fmts = [None] + [NUM_FMT_4] * 15
    for out_r, i in enumerate(order, start=2):
        row = [
            g0.labels[i],
            g0.in_degree.mean[i],   g0.in_degree.ci_low[i],   g0.in_degree.ci_high[i],
            g0.out_degree.mean[i],  g0.out_degree.ci_low[i],  g0.out_degree.ci_high[i],
            g0.total_degree.mean[i],g0.total_degree.ci_low[i],g0.total_degree.ci_high[i],
            g0.eigen_centr.mean[i], g0.eigen_centr.ci_low[i], g0.eigen_centr.ci_high[i],
            g0.betweenness.mean[i], g0.betweenness.ci_low[i], g0.betweenness.ci_high[i],
        ]
        _write_row(ws, out_r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 4 — G0 EDGE PROBABILITIES (ADMISSIBLE EDGES ONLY)
# ============================================================

def _build_G0_edges(ws: Worksheet, edge_bundle: EdgeAnalysisBundle) -> None:
    headers = ["source", "target", "edge_prob", "ci_lower", "n_eff",
               "selected_naive", "selected_credible"]
    _write_header(ws, 1, headers)

    ea = edge_bundle.G0
    P = ea.edge_prob
    ny = P.shape[0]

    # Extract all admissible cells, sort by edge_prob descending
    rows = []
    for i in range(ny):
        for j in range(ny):
            if ea.admissibility[i, j]:
                rows.append((
                    ea.col_labels[j],            # source (parent)
                    ea.row_labels[i],            # target (response)
                    float(P[i, j]),
                    float(ea.ci_lower[i, j]),
                    float(ea.n_eff[i, j]) if not np.isnan(ea.n_eff[i, j]) else None,
                    bool(ea.selected_naive[i, j]),
                    bool(ea.selected_credible[i, j]),
                ))
    rows.sort(key=lambda r: -r[2])

    fmts = [None, None, NUM_FMT_4, NUM_FMT_4, NUM_FMT_INT, None, None]
    for out_r, row in enumerate(rows, start=2):
        _write_row(ws, out_r, list(row), fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 5 — PHI NODE METRICS (LAG1, LAG7, AGG)
# ============================================================

def _build_Phi_nodes(ws: Worksheet,
                     phi_results: dict[str, WeightedMetricsBundle],
                     labels: list[str]) -> None:
    # Wide format: one row per country, columns grouped by slice
    slices = list(phi_results.keys())   # typically ["lag1", "lag7", "agg"]

    headers = ["country"]
    for s in slices:
        headers += [
            f"{s}_in_str", f"{s}_out_str", f"{s}_tot_str",
            f"{s}_eigen",  f"{s}_between",
        ]
    _write_header(ws, 1, headers)

    # Sort countries by aggregate total_strength desc (best summary)
    if "agg" in phi_results:
        sort_vals = phi_results["agg"].total_strength.values
    else:
        sort_vals = phi_results[slices[0]].total_strength.values
    order = np.argsort(-sort_vals, kind="stable")

    fmts = [None] + [NUM_FMT_4] * (5 * len(slices))
    for out_r, i in enumerate(order, start=2):
        row = [labels[i]]
        for s in slices:
            b = phi_results[s]
            row += [
                b.in_strength.values[i],
                b.out_strength.values[i],
                b.total_strength.values[i],
                b.eigen_centr.values[i],
                b.betweenness.values[i],
            ]
        _write_row(ws, out_r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 6 — PHI GRAPH-LEVEL
# ============================================================

def _build_Phi_graph_level(ws: Worksheet,
                           phi_results: dict[str, WeightedMetricsBundle]) -> None:
    headers = ["slice", "total_weight", "mean_weight_off_diag",
               "weight_density", "weight_gini"]
    _write_header(ws, 1, headers)

    fmts = [None, NUM_FMT_4, NUM_FMT_4, NUM_FMT_4, NUM_FMT_4]
    for r, (slice_name, b) in enumerate(phi_results.items(), start=2):
        row = [
            slice_name,
            b.total_weight.value,
            b.mean_weight.value,
            b.weight_density.value,
            b.weight_gini.value,
        ]
        _write_row(ws, r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 7 — GAMMA EXPOSURE
# ============================================================

def _build_Gamma_exposure(ws: Worksheet, gamma_res: GammaAnalysisResults) -> None:
    headers = ["country", "total", "total_wind", "total_solar",
               "own_wind", "own_solar", "cross_wind", "cross_solar",
               "share_own_pct", "share_wind_pct"]
    _write_header(ws, 1, headers)

    expo = gamma_res.exposure_summary
    order = np.argsort(-expo.total, kind="stable")

    fmts = [None] + [NUM_FMT_4] * 7 + [NUM_FMT_3, NUM_FMT_3]
    for out_r, i in enumerate(order, start=2):
        row = [
            expo.labels[i],
            expo.total[i],
            expo.total_wind[i],
            expo.total_solar[i],
            expo.own_wind[i],
            expo.own_solar[i],
            expo.cross_wind[i],
            expo.cross_solar[i],
            100.0 * expo.share_own[i],
            100.0 * expo.share_wind[i],
        ]
        _write_row(ws, out_r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 8 — GAMMA PERVASIVENESS
# ============================================================

def _build_Gamma_pervasiveness(ws: Worksheet,
                               gamma_res: GammaAnalysisResults,
                               country_labels: list[str]) -> None:
    headers = ["exog", "source_country", "type", "out_strength", "n_targets", "rank"]
    _write_header(ws, 1, headers)

    perv = gamma_res.pervasiveness
    order = np.argsort(-perv.out_strength, kind="stable")

    fmts = [None, None, None, NUM_FMT_4, NUM_FMT_INT, NUM_FMT_INT]
    for out_r, j in enumerate(order, start=2):
        row = [
            perv.labels[j],
            country_labels[perv.source_country[j]],
            "wind" if perv.is_wind[j] else "solar",
            float(perv.out_strength[j]),
            int(perv.n_targets[j]),
            out_r - 1,    # rank (1-based, by sorted position)
        ]
        _write_row(ws, out_r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 9 — GAMMA COUNTRY-COLLAPSED NETWORK NODES
# ============================================================

def _build_Gamma_country_net_nodes(ws: Worksheet,
                                   gamma_res: GammaAnalysisResults) -> None:
    # Three networks side by side: wind, solar, combined.
    # Per-country: in_str, out_str, tot_str, eigen, between, for each.

    bundles = {
        "wind":     gamma_res.country_network.wind_bundle,
        "solar":    gamma_res.country_network.solar_bundle,
        "combined": gamma_res.country_network.combined_bundle,
    }
    headers = ["country"]
    for s in bundles:
        headers += [f"{s}_in_str", f"{s}_out_str", f"{s}_tot_str",
                    f"{s}_eigen",  f"{s}_between"]
    _write_header(ws, 1, headers)

    # Sort by combined total_strength descending
    combined = bundles["combined"]
    sort_vals = combined.total_strength.values
    order = np.argsort(-sort_vals, kind="stable")
    labels = combined.labels

    fmts = [None] + [NUM_FMT_4] * (5 * len(bundles))
    for out_r, i in enumerate(order, start=2):
        row = [labels[i]]
        for s, b in bundles.items():
            row += [
                b.in_strength.values[i],
                b.out_strength.values[i],
                b.total_strength.values[i],
                b.eigen_centr.values[i],
                b.betweenness.values[i],
            ]
        _write_row(ws, out_r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 10 — GAMMA COUNTRY-COLLAPSED NETWORK GRAPH-LEVEL
# ============================================================

def _build_Gamma_country_net_graph(ws: Worksheet,
                                   gamma_res: GammaAnalysisResults) -> None:
    headers = ["network", "total_weight", "mean_weight_off_diag",
               "weight_density", "weight_gini"]
    _write_header(ws, 1, headers)

    bundles = [
        ("wind",     gamma_res.country_network.wind_bundle),
        ("solar",    gamma_res.country_network.solar_bundle),
        ("combined", gamma_res.country_network.combined_bundle),
    ]

    fmts = [None, NUM_FMT_4, NUM_FMT_4, NUM_FMT_4, NUM_FMT_4]
    for r, (name, b) in enumerate(bundles, start=2):
        row = [
            name,
            b.total_weight.value,
            b.mean_weight.value,
            b.weight_density.value,
            b.weight_gini.value,
        ]
        _write_row(ws, r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# SHEET 11 — CROSS COMPARISON
# ============================================================

def _build_cross_comparison(ws: Worksheet,
                            g0: G0MetricsBundle,
                            phi_agg: WeightedMetricsBundle,
                            gamma_res: GammaAnalysisResults) -> None:
    """
    For each country, report:
      - G0      : out_degree mean and rank
      - Phi-agg : out_strength and rank
      - Gamma   : combined out_strength and rank
    Sorted by G0 rank (top hubs in G0 come first), so the reader sees
    immediately which paesi are hubs in ALL three networks vs only in one.
    """
    headers = ["country",
               "G0_out_deg",     "G0_rank",
               "Phi_out_str",    "Phi_rank",
               "Gamma_out_str",  "Gamma_rank",
               "max_rank",       "min_rank"]
    _write_header(ws, 1, headers)

    g0_vals   = g0.out_degree.mean
    phi_vals  = phi_agg.out_strength.values
    gam_vals  = gamma_res.country_network.combined_bundle.out_strength.values

    g0_rank   = _rank_descending(g0_vals)
    phi_rank  = _rank_descending(phi_vals)
    gam_rank  = _rank_descending(gam_vals)

    max_rank  = np.maximum.reduce([g0_rank, phi_rank, gam_rank])
    min_rank  = np.minimum.reduce([g0_rank, phi_rank, gam_rank])

    # Sort by best (lowest) min_rank first, then by g0_rank as tie-breaker
    order = np.lexsort((g0_rank, min_rank))

    fmts = [None, NUM_FMT_4, NUM_FMT_INT, NUM_FMT_4, NUM_FMT_INT,
            NUM_FMT_4, NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT]
    for out_r, i in enumerate(order, start=2):
        row = [
            g0.labels[i],
            float(g0_vals[i]),  int(g0_rank[i]),
            float(phi_vals[i]), int(phi_rank[i]),
            float(gam_vals[i]), int(gam_rank[i]),
            int(max_rank[i]),
            int(min_rank[i]),
        ]
        _write_row(ws, out_r, row, fmts)

    _freeze_header(ws)
    _autosize_columns(ws)


# ============================================================
# TOP-LEVEL
# ============================================================

def export_to_excel(
    chain:        ChainOutput,
    edge_bundle:  EdgeAnalysisBundle,
    g0_metrics:   G0MetricsBundle,
    phi_results:  dict[str, WeightedMetricsBundle],
    gamma_res:    GammaAnalysisResults,
    output_path:  str | Path = "network_metrics.xlsx",
    alpha:        float = 0.10,
) -> Path:
    """
    Write all results to a multi-sheet Excel workbook.

    Returns the path of the saved file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    print(f"[excel] Writing workbook to {output_path.resolve()}")

    ws = wb.create_sheet("README")
    _build_readme(ws, chain, alpha=alpha)
    print(f"  + README")

    ws = wb.create_sheet("G0_summary")
    _build_G0_summary(ws, g0_metrics)
    print(f"  + G0_summary")

    ws = wb.create_sheet("G0_nodes")
    _build_G0_nodes(ws, g0_metrics)
    print(f"  + G0_nodes")

    ws = wb.create_sheet("G0_edges")
    _build_G0_edges(ws, edge_bundle)
    print(f"  + G0_edges")

    ws = wb.create_sheet("Phi_nodes")
    _build_Phi_nodes(ws, phi_results, chain.country_labels)
    print(f"  + Phi_nodes")

    ws = wb.create_sheet("Phi_graph_level")
    _build_Phi_graph_level(ws, phi_results)
    print(f"  + Phi_graph_level")

    ws = wb.create_sheet("Gamma_exposure")
    _build_Gamma_exposure(ws, gamma_res)
    print(f"  + Gamma_exposure")

    ws = wb.create_sheet("Gamma_pervasiveness")
    _build_Gamma_pervasiveness(ws, gamma_res, chain.country_labels)
    print(f"  + Gamma_pervasiveness")

    ws = wb.create_sheet("Gamma_country_net_nodes")
    _build_Gamma_country_net_nodes(ws, gamma_res)
    print(f"  + Gamma_country_net_nodes")

    ws = wb.create_sheet("Gamma_country_net_graph")
    _build_Gamma_country_net_graph(ws, gamma_res)
    print(f"  + Gamma_country_net_graph")

    ws = wb.create_sheet("Cross_comparison")
    _build_cross_comparison(ws, g0_metrics, phi_results["agg"], gamma_res)
    print(f"  + Cross_comparison")

    wb.save(output_path)
    print(f"[excel] Done.")
    return output_path


